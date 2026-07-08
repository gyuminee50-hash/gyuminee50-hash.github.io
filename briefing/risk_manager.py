"""
리스크관리팀 — AI 패밀리오피스 ①
==========================================
매일 07:00 실행
- 투자일지 엑셀에서 포지션 읽기 (미래에셋 + 메리츠)
- yfinance로 메리츠 달러 종목 현재가 조회
- 섹터별 비중 계산
- Groq 70B 리스크 분석 (편중 경고 + 권고)
- 텔레그램 발송
"""
import json, os, sys
import openpyxl
import yfinance as yf
from datetime import datetime
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

import groq_client
from fo_utils import save_status, send as tg_send

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = r'C:\Users\DeskTop\OneDrive\문서\GMCapital_투자일지.xlsx'

with open(os.path.join(BASE_DIR, 'config.json'), 'r', encoding='utf-8') as f:
    _cfg = json.load(f)

# 환율 (메리츠 달러 종목 → 원화 환산용)
def _get_usdkrw():
    try:
        t = yf.Ticker('USDKRW=X')
        return float(t.fast_info.last_price)
    except Exception:
        return 1380.0  # 조회 실패 시 기본값

# 섹터 분류표 (티커 → 섹터)
SECTOR_MAP = {
    # 미래에셋 (한국 ETF)
    '133690': '미국기술주ETF',
    '360750': '미국시장ETF',
    '232080': '한국성장주ETF',
    # 메리츠 (미국 개별주)
    'TSM':  '반도체',
    'MU':   '반도체',
    'GGLL': '빅테크레버리지',
    'IEMG': '신흥국ETF',
    'SPY':  '미국시장ETF',
    'QQQ':  '미국기술주ETF',
    # 기타 공통
    'NVDA': '반도체',
    'AMD':  '반도체',
    'AVGO': '반도체',
    'MSFT': '빅테크',
    'META': '빅테크',
    'AMZN': '빅테크',
    'GOOG': '빅테크',
    'GOOGL':'빅테크',
}

def _sector(ticker):
    return SECTOR_MAP.get(str(ticker).upper(), '기타')


def _read_miraeasset_positions(wb):
    """미래에셋_포지션 시트 → 포지션 리스트 (원화)."""
    ws = wb['미래에셋_포지션']
    positions = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row and str(row[0]).strip() == '티커':
                header_found = True
            continue
        if not row or row[0] is None:
            continue
        ticker = str(row[0]).strip()
        name   = str(row[1]).strip() if row[1] else ''
        qty    = row[4] or 0     # 보유수량
        price  = row[7] or 0     # 현재가(₩)
        val    = row[8] or 0     # 평가금액(₩)
        if qty == 0:
            continue
        positions.append({
            'ticker': ticker, 'name': name,
            'qty': qty, 'price_krw': price,
            'value_krw': val,
            'account': '미래에셋',
            'currency': 'KRW',
            'sector': _sector(ticker),
        })
    return positions


def _read_meritz_positions(wb, usdkrw):
    """메리츠_거래 내역 → 순 보유수량 계산 → yfinance 현재가 → 평가액(원화 환산)."""
    ws = wb['메리츠_거래']
    holdings = defaultdict(lambda: {'name': '', 'qty': 0, 'avg_rate': 1380.0})
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row and str(row[0]).strip() == '날짜':
                header_found = True
            continue
        if not row or row[0] is None:
            continue
        구분 = str(row[1]).strip() if row[1] else ''
        ticker = str(row[2]).strip().upper() if row[2] else ''
        name   = str(row[3]).strip() if row[3] else ''
        qty    = float(row[5]) if row[5] else 0
        rate   = float(row[6]) if row[6] else usdkrw

        if not ticker:
            continue
        if 구분 == '매수':
            holdings[ticker]['qty']  += qty
            holdings[ticker]['name']  = name
            holdings[ticker]['avg_rate'] = rate
        elif 구분 == '매도':
            holdings[ticker]['qty']  -= qty

    positions = []
    for ticker, info in holdings.items():
        if info['qty'] <= 0:
            continue
        # yfinance 현재가 조회
        try:
            price_usd = float(yf.Ticker(ticker).fast_info.last_price)
        except Exception:
            price_usd = 0.0
        val_krw = price_usd * info['qty'] * usdkrw
        positions.append({
            'ticker': ticker,
            'name': info['name'],
            'qty': info['qty'],
            'price_usd': round(price_usd, 2),
            'value_krw': round(val_krw),
            'account': '메리츠',
            'currency': 'USD',
            'sector': _sector(ticker),
        })
    return positions


def _sector_summary(positions):
    """섹터별 합계 + 비중 계산."""
    total = sum(p['value_krw'] for p in positions)
    by_sector = defaultdict(float)
    for p in positions:
        by_sector[p['sector']] += p['value_krw']
    result = sorted(
        [{'sector': s, 'value': v, 'pct': round(v / total * 100, 1) if total else 0}
         for s, v in by_sector.items()],
        key=lambda x: x['value'], reverse=True,
    )
    return result, total


_RISK_PROMPT = """\
당신은 GM Capital의 리스크관리팀장입니다.
오늘 날짜: {today}

현재 포트폴리오:
{portfolio_str}

섹터별 비중:
{sector_str}

총 평가액: {total_krw:,}원 (약 ${total_usd:,.0f})

아래 형식으로 리스크 리포트를 작성하라:

[리스크 수준] 낮음 / 보통 / 높음
[편중 경고] (섹터 비중 50% 초과 시 경고 / 없으면 "없음")
[주요 리스크] 1~3개 항목
[권고] 한 줄

원칙:
- 섹터 비중 50% 초과 = 편중 경고
- 레버리지 ETF(GGLL 등) 보유는 변동성 리스크로 별도 언급
- 억지 경고 금지 — 실제 리스크만"""

def _groq_risk_analysis(positions, sector_summary, total_krw, usdkrw):
    portfolio_lines = '\n'.join(
        f"  {p['account']} | {p['ticker']} ({p['name']}) | "
        f"수량 {p['qty']} | 평가액 {p['value_krw']:,}원"
        for p in positions
    )
    sector_lines = '\n'.join(
        f"  {s['sector']}: {s['pct']}% ({s['value']:,}원)"
        for s in sector_summary
    )
    total_usd = total_krw / usdkrw
    try:
        return groq_client.call(
            _RISK_PROMPT.format(
                today=datetime.now().strftime('%Y-%m-%d'),
                portfolio_str=portfolio_lines,
                sector_str=sector_lines,
                total_krw=int(total_krw),
                total_usd=total_usd,
            ),
            max_tokens=400, temperature=0.2,
            model='llama-3.3-70b-versatile',
        )
    except Exception as e:
        return f'[리스크 분석 오류] {e}'


def run_risk_check():
    print(f'[{datetime.now().strftime("%Y-%m-%d %H:%M")}] 리스크관리팀 점검 시작...')

    usdkrw = _get_usdkrw()
    print(f'  USD/KRW: {usdkrw:.0f}')

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except Exception as e:
        print(f'  [엑셀 오류] {e}')
        return

    positions  = _read_miraeasset_positions(wb)
    positions += _read_meritz_positions(wb, usdkrw)
    wb.close()

    if not positions:
        print('  포지션 없음 — 종료')
        return

    sector_sum, total_krw = _sector_summary(positions)
    print(f'  포지션 {len(positions)}종목, 총 {total_krw:,}원')

    analysis = _groq_risk_analysis(positions, sector_sum, total_krw, usdkrw)

    warnings  = [s for s in sector_sum if s['pct'] >= 50]
    warn_icon = '⚠️' if warnings else '✅'

    # 리스크 수준 파싱
    risk_level = '보통'
    for line in analysis.splitlines():
        if '[리스크 수준]' in line or '리스크 수준:' in line:
            risk_level = line.split()[-1]
            break

    now_str = datetime.now().strftime('%m/%d %H:%M')

    # 섹터 비중 한 줄 압축
    sector_line = ' | '.join(
        f"{'🔴' if s['pct']>=50 else ('🟡' if s['pct']>=30 else '🟢')}{s['sector']} {s['pct']}%"
        for s in sector_sum
    )

    msg = '\n'.join([
        f'<b>{warn_icon} 리스크관리팀  {now_str}  |  리스크 {risk_level}</b>',
        f'<b>💰 {int(total_krw):,}원</b>  <i>(환율 {usdkrw:.0f})</i>',
        '',
        sector_line,
        '',
        analysis,
    ])

    tg_send(msg)

    save_status('risk', {
        'risk_level': risk_level,
        'total_krw':  int(total_krw),
        'warnings':   [s['sector'] for s in warnings],
        'sector_top': sector_sum[0] if sector_sum else {},
    })
    print('✅ 리스크 리포트 발송 완료!')


if __name__ == '__main__':
    run_risk_check()
