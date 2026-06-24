"""
실수감시팀 — AI 패밀리오피스 ③
==========================================
매일 08:30 실행
- FOMO 감지: 최근 14일 내 매수 종목이 매수 전 30일 20%+ 급등했으면 경고
- 손절회피 감지: 현재 손실 -25% 이하 + 보유 중 → 최초 논리 재점검 요구
- Groq 70B 심리 패턴 분석
- 문제 감지 시에만 텔레그램 발송 (0건 정상)
"""
import json, os, sys
import openpyxl
import yfinance as yf
from datetime import datetime, timedelta, date
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

import groq_client
from fo_utils import save_status, send as tg_send

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = r'C:\Users\DeskTop\OneDrive\문서\GMCapital_투자일지.xlsx'

with open(os.path.join(BASE_DIR, 'config.json'), 'r', encoding='utf-8') as f:
    _cfg = json.load(f)

FOMO_LOOKBACK_DAYS  = 14   # 최근 N일 내 매수 거래 체크
FOMO_SURGE_PCT      = 20.0 # 매수 전 30일 N% 이상 급등이면 FOMO 의심
STOPLOSS_THRESHOLD  = -25.0  # 손실 N% 이하 → 손절회피 경고

_FOMO_PROMPT = """\
투자자가 {ticker}({name})를 {buy_date}에 매수했다.
매수 전 30일 동안 이 종목은 {pre_surge:+.1f}% 상승했었다.
현재 매수 후 수익률: {cur_ret:+.1f}%

이 매수가 FOMO(감정적 추격매수)일 가능성을 분석하라.
형식:
FOMO 가능성: 높음 / 보통 / 낮음
근거: 한 줄
권고: 한 줄

원칙: 확신 없으면 낮음 — 억지 경고 금지"""

_STOPLOSS_PROMPT = """\
투자자가 {ticker}({name})에서 현재 {ret:+.1f}% 손실 중이며 계속 보유 중이다.
최초 매수 논거: {reasoning}
보유 기간: {hold_days}일

손절회피(손실 인정 회피) 패턴인지 분석하라.
형식:
손절회피 위험: 높음 / 보통 / 낮음
근거: 한 줄
권고: 한 줄 (손절기준가 제시)

원칙: 최초 매수 논거가 현재도 유효하면 낮음으로 판단"""


def _get_usdkrw():
    try:
        return float(yf.Ticker('USDKRW=X').fast_info.last_price)
    except Exception:
        return 1380.0


def _pre_surge(ticker, buy_date_str, currency='USD'):
    """매수일 기준 30일 전~매수일 주가 변동률."""
    try:
        buy_dt  = datetime.strptime(buy_date_str[:10], '%Y-%m-%d').date()
        end_dt  = buy_dt
        start_dt= buy_dt - timedelta(days=35)
        # 한국 ETF는 yfinance 티커 형식 다름
        yfticker = ticker if currency == 'USD' else f'{ticker}.KS'
        df = yf.download(yfticker, start=str(start_dt), end=str(end_dt + timedelta(days=1)),
                         interval='1d', auto_adjust=True, progress=False)
        if df.empty or len(df) < 5:
            return None
        start_px = float(df['Close'].iloc[0])
        end_px   = float(df['Close'].iloc[-1])
        return (end_px - start_px) / start_px * 100
    except Exception:
        return None


def _read_recent_buys(wb, days=14):
    """최근 N일 내 매수 거래 추출 (미래에셋 + 메리츠)."""
    cutoff = date.today() - timedelta(days=days)
    buys   = []

    for sheet_name, currency in [('미래에셋_거래', 'KRW'), ('메리츠_거래', 'USD')]:
        ws = wb[sheet_name]
        header_found = False
        for row in ws.iter_rows(values_only=True):
            if not header_found:
                if row and str(row[0]).strip() == '날짜':
                    header_found = True
                continue
            if not row or row[0] is None:
                continue
            try:
                trade_date = datetime.strptime(str(row[0]).strip()[:10], '%Y-%m-%d').date()
            except Exception:
                continue
            구분 = str(row[1]).strip() if row[1] else ''
            if 구분 != '매수' or trade_date < cutoff:
                continue
            buys.append({
                'date':     str(row[0]).strip()[:10],
                'ticker':   str(row[2]).strip().upper() if row[2] else '',
                'name':     str(row[3]).strip() if row[3] else '',
                'currency': currency,
                'account':  sheet_name.split('_')[0],
            })
    return buys


def _read_positions(wb, usdkrw):
    """손절회피 체크용 포지션 (수익률 포함)."""
    positions = []

    # 미래에셋_포지션
    ws = wb['미래에셋_포지션']
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row and str(row[0]).strip() == '티커':
                header_found = True
            continue
        if not row or row[0] is None:
            continue
        qty = row[4] or 0
        if qty == 0:
            continue
        positions.append({
            'ticker':    str(row[0]).strip(),
            'name':      str(row[1]).strip() if row[1] else '',
            'ret':       float(row[10]) if row[10] else 0.0,
            'reasoning': str(row[11]).strip()[:120] if row[11] else '기록 없음',
            'currency':  'KRW',
            'buy_date':  None,
        })

    # 메리츠 — 거래내역에서 최초 매수일 + 평균단가 + 현재수익률
    ws2 = wb['메리츠_거래']
    holdings = defaultdict(lambda: {'name':'','qty':0,'invest':0.0,'first_date':None,'reasoning':'기록 없음'})
    header_found = False
    for row in ws2.iter_rows(values_only=True):
        if not header_found:
            if row and str(row[0]).strip() == '날짜':
                header_found = True
            continue
        if not row or row[0] is None:
            continue
        구분 = str(row[1]).strip() if row[1] else ''
        ticker = str(row[2]).strip().upper() if row[2] else ''
        name   = str(row[3]).strip() if row[3] else ''
        price  = float(row[4]) if row[4] else 0.0
        qty    = float(row[5]) if row[5] else 0.0
        tdate  = str(row[0]).strip()[:10] if row[0] else ''
        if not ticker:
            continue
        if 구분 == '매수':
            holdings[ticker]['qty'] += qty
            holdings[ticker]['name'] = name
            holdings[ticker]['invest'] += price * qty
            if holdings[ticker]['first_date'] is None:
                holdings[ticker]['first_date'] = tdate
        elif 구분 == '매도':
            holdings[ticker]['qty'] -= qty

    for ticker, info in holdings.items():
        if info['qty'] <= 0:
            continue
        try:
            price_usd = float(yf.Ticker(ticker).fast_info.last_price)
        except Exception:
            continue
        avg_usd = info['invest'] / info['qty'] if info['qty'] else 0
        ret     = (price_usd - avg_usd) / avg_usd * 100 if avg_usd else 0
        # 보유 기간
        hold_days = 0
        if info['first_date']:
            try:
                fd = datetime.strptime(info['first_date'], '%Y-%m-%d').date()
                hold_days = (date.today() - fd).days
            except Exception:
                pass
        positions.append({
            'ticker':    ticker,
            'name':      info['name'],
            'ret':       round(ret, 2),
            'reasoning': info['reasoning'],
            'currency':  'USD',
            'buy_date':  info['first_date'],
            'hold_days': hold_days,
        })
    return positions


def run_mistake_monitor():
    print(f'[{datetime.now().strftime("%Y-%m-%d %H:%M")}] 실수감시팀 점검 시작...')

    usdkrw = _get_usdkrw()
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except Exception as e:
        print(f'  [엑셀 오류] {e}')
        return

    recent_buys = _read_recent_buys(wb, FOMO_LOOKBACK_DAYS)
    positions   = _read_positions(wb, usdkrw)
    wb.close()

    alerts = []

    # ── FOMO 감지 ────────────────────────────────────────────────────
    for buy in recent_buys:
        ticker = buy['ticker']
        if not ticker:
            continue
        print(f'  [FOMO 체크] {ticker}...')
        surge = _pre_surge(ticker, buy['date'], buy['currency'])
        if surge is None or surge < FOMO_SURGE_PCT:
            print(f'    → 급등 {surge:.1f}% < {FOMO_SURGE_PCT}% — 해당 없음' if surge is not None else '    → 데이터 없음')
            continue

        # 현재 수익률
        cur_ret = next((p['ret'] for p in positions if p['ticker'] == ticker), 0.0)
        try:
            analysis = groq_client.call(
                _FOMO_PROMPT.format(
                    ticker=ticker, name=buy['name'],
                    buy_date=buy['date'], pre_surge=surge, cur_ret=cur_ret,
                ),
                max_tokens=150, temperature=0.2,
                model='llama-3.3-70b-versatile',
            )
        except Exception as e:
            analysis = f'분석 오류: {e}'

        lines = {l.split(':')[0].strip(): ':'.join(l.split(':')[1:]).strip()
                 for l in analysis.strip().splitlines() if ':' in l}
        level = lines.get('FOMO 가능성', '낮음')
        if level in ('높음', '보통'):
            alerts.append({
                'type': 'FOMO',
                'ticker': ticker,
                'name': buy['name'],
                'detail': f'매수 전 30일 {surge:+.1f}% 급등 후 매수',
                'analysis': analysis,
                'level': level,
            })
            print(f'    → FOMO 가능성 {level} !')

    # ── 손절회피 감지 ─────────────────────────────────────────────────
    for pos in positions:
        if pos['ret'] > STOPLOSS_THRESHOLD:
            continue
        print(f'  [손절회피 체크] {pos["ticker"]} ({pos["ret"]:+.1f}%)...')
        hold_days = pos.get('hold_days', 0)
        try:
            analysis = groq_client.call(
                _STOPLOSS_PROMPT.format(
                    ticker=pos['ticker'], name=pos['name'],
                    ret=pos['ret'], reasoning=pos['reasoning'],
                    hold_days=hold_days,
                ),
                max_tokens=150, temperature=0.2,
                model='llama-3.3-70b-versatile',
            )
        except Exception as e:
            analysis = f'분석 오류: {e}'

        lines = {l.split(':')[0].strip(): ':'.join(l.split(':')[1:]).strip()
                 for l in analysis.strip().splitlines() if ':' in l}
        level = lines.get('손절회피 위험', '낮음')
        if level in ('높음', '보통'):
            alerts.append({
                'type': '손절회피',
                'ticker': pos['ticker'],
                'name': pos['name'],
                'detail': f'손실 {pos["ret"]:+.1f}%  보유 {hold_days}일',
                'analysis': analysis,
                'level': level,
            })
            print(f'    → 손절회피 위험 {level} !')

    save_status('mistakes', {'alert_count': len(alerts),
                              'tickers': [a['ticker'] for a in alerts]})

    if not alerts:
        print('✅ 실수 패턴 없음 (0건 정상)')
        return

    now_str = datetime.now().strftime('%m/%d %H:%M')
    lines   = [f'<b>⚠️ 실수감시팀  {now_str}  |  경고 {len(alerts)}건</b>']
    for a in alerts:
        icon = '🔴' if a['level'] == '높음' else '🟡'
        lines.append(f'\n{icon} <b>[{a["type"]}] {a["ticker"]}</b>  {a["detail"]}')
        lines.append(a['analysis'])
    lines.append('\n<i>* 참고용 — CEO 최종 판단 우선. 0건이 정상.</i>')

    tg_send('\n'.join(lines))
    print(f'✅ 실수 경고 {len(alerts)}건 발송!')


if __name__ == '__main__':
    run_mistake_monitor()
