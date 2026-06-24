"""
AI 투자위원회 — AI 패밀리오피스 ②
==========================================
매주 월요일 08:00 실행
- 투자일지 엑셀에서 보유 종목 + 수익률 읽기
- 종목별 Bull/Bear/중립 토론 (Groq 70B, 하나의 프롬프트)
- 근거 없으면 "뚜렷한 반론 없음" — 억지 토론 금지
- CEO(규민님) 최종 판단용 텔레그램 발송
"""
import json, os, sys
import openpyxl
import requests
import yfinance as yf
from datetime import datetime
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

import groq_client

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = r'C:\Users\DeskTop\OneDrive\문서\GMCapital_투자일지.xlsx'

with open(os.path.join(BASE_DIR, 'config.json'), 'r', encoding='utf-8') as f:
    _cfg = json.load(f)

_COMMITTEE_PROMPT = """\
당신은 GM Capital의 AI 투자위원회입니다.
오늘: {today}

종목: {ticker} ({name})
현재가: {price}
평균단가: {avg_price}
수익률: {ret:+.1f}%
투자원금: {invest:,}원
평가금액: {value:,}원
매수 논거: {reasoning}

아래 형식으로 정확히 작성하라. 다른 텍스트 없이:

[Bull] (낙관 근거 2~3줄. 없으면 "뚜렷한 낙관 근거 없음")
[Bear] (비관 근거 2~3줄. 없으면 "뚜렷한 반론 없음")
[중립] (양측 종합 균형 판단 1~2줄)
[권고] 매수확대 / 유지 / 축소검토 / 매도검토

핵심 원칙:
- 근거 없으면 없다고 함 — 억지 반대 금지
- 모르면 모른다고 함
- 현재 수익률·시장 상황 기반으로만 판단"""


def _read_positions(wb, usdkrw):
    """미래에셋 + 메리츠 포지션 읽기."""
    positions = []

    # 미래에셋_포지션
    ws = wb['미래에셋_포지션']
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row and str(row[0]).strip() == '티커':
                header_found = True
            continue
        if not row or row[0] is None:
            continue
        qty = row[4] or 0
        if qty == 0:
            continue
        ret = float(row[10]) if row[10] else 0.0
        positions.append({
            'ticker':    str(row[0]).strip(),
            'name':      str(row[1]).strip() if row[1] else '',
            'qty':       qty,
            'avg_price': row[5] or 0,
            'invest':    int(row[6] or 0),
            'price':     row[7] or 0,
            'value':     int(row[8] or 0),
            'ret':       ret,
            'reasoning': str(row[11]).strip()[:100] if row[11] else '기록 없음',
            'account':   '미래에셋',
            'currency':  'KRW',
        })

    # 메리츠 — 거래내역에서 순 보유수량 계산
    ws2 = wb['메리츠_거래']
    holdings = defaultdict(lambda: {'name':'','qty':0,'invest_usd':0.0,'avg_usd':0.0,'reasoning':'기록 없음'})
    header_found = False
    for row in ws2.iter_rows(values_only=True):
        if not header_found:
            if row and str(row[0]).strip() == '날짜':
                header_found = True
            continue
        if not row or row[0] is None:
            continue
        구분 = str(row[1]).strip() if row[1] else ''
        ticker = str(row[2]).strip().upper() if row[2] else ''
        name   = str(row[3]).strip() if row[3] else ''
        price  = float(row[4]) if row[4] else 0.0
        qty    = float(row[5]) if row[5] else 0.0
        if not ticker:
            continue
        if 구분 == '매수':
            holdings[ticker]['qty'] += qty
            holdings[ticker]['name'] = name
            holdings[ticker]['invest_usd'] += price * qty
        elif 구분 == '매도':
            holdings[ticker]['qty'] -= qty

    for ticker, info in holdings.items():
        if info['qty'] <= 0:
            continue
        try:
            price_usd = float(yf.Ticker(ticker).fast_info.last_price)
        except Exception:
            price_usd = 0.0
        avg_usd   = info['invest_usd'] / info['qty'] if info['qty'] else 0
        value_krw = int(price_usd * info['qty'] * usdkrw)
        invest_krw= int(info['invest_usd'] * usdkrw)
        ret = ((price_usd - avg_usd) / avg_usd * 100) if avg_usd else 0.0
        positions.append({
            'ticker':    ticker,
            'name':      info['name'],
            'qty':       info['qty'],
            'avg_price': round(avg_usd, 2),
            'invest':    invest_krw,
            'price':     round(price_usd, 2),
            'value':     value_krw,
            'ret':       round(ret, 2),
            'reasoning': info['reasoning'],
            'account':   '메리츠',
            'currency':  'USD',
        })

    return positions


def _get_usdkrw():
    try:
        return float(yf.Ticker('USDKRW=X').fast_info.last_price)
    except Exception:
        return 1380.0


def _debate(pos):
    """종목 하나에 대해 Bull/Bear/중립 토론 생성."""
    price_str = (f"${pos['price']}" if pos['currency'] == 'USD'
                 else f"₩{pos['price']:,}")
    avg_str   = (f"${pos['avg_price']}" if pos['currency'] == 'USD'
                 else f"₩{pos['avg_price']:,}")
    try:
        return groq_client.call(
            _COMMITTEE_PROMPT.format(
                today=datetime.now().strftime('%Y-%m-%d'),
                ticker=pos['ticker'],
                name=pos['name'],
                price=price_str,
                avg_price=avg_str,
                ret=pos['ret'],
                invest=pos['invest'],
                value=pos['value'],
                reasoning=pos['reasoning'],
            ),
            max_tokens=350, temperature=0.3,
            model='llama-3.3-70b-versatile',
        )
    except Exception as e:
        return f'[토론 생성 오류] {e}'


def _send_telegram(text):
    token   = _cfg['telegram_token']
    chat_id = _cfg['telegram_chat_id']
    # 텔레그램 4096자 제한 → 분할 발송
    for i in range(0, len(text), 4000):
        requests.post(
            f'https://api.telegram.org/bot{token}/sendMessage',
            json={'chat_id': chat_id, 'text': text[i:i+4000], 'parse_mode': 'HTML'},
            timeout=15,
        )


def run_committee():
    print(f'[{datetime.now().strftime("%Y-%m-%d %H:%M")}] AI 투자위원회 시작...')

    usdkrw = _get_usdkrw()
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except Exception as e:
        print(f'  [엑셀 오류] {e}')
        return

    positions = _read_positions(wb, usdkrw)
    wb.close()

    if not positions:
        print('  포지션 없음 — 종료')
        return

    total_value = sum(p['value'] for p in positions)
    total_invest = sum(p['invest'] for p in positions)
    total_ret = (total_value - total_invest) / total_invest * 100 if total_invest else 0

    now_str = datetime.now().strftime('%m/%d %H:%M')
    header  = (
        f'<b>🏛 AI 투자위원회  {now_str}</b>\n'
        f'총 평가액 {total_value:,}원 | 총 수익률 {total_ret:+.1f}%\n'
        f'보유 {len(positions)}종목 — 아래 CEO 최종 판단 요청\n'
    )
    _send_telegram(header)

    for pos in positions:
        print(f'  [{pos["ticker"]}] 토론 중...')
        debate = _debate(pos)

        ret_icon = '📈' if pos['ret'] >= 0 else '📉'
        msg = (
            f'<b>{ret_icon} {pos["ticker"]} ({pos["name"]})</b>\n'
            f'현재가 {pos["price"]}  |  수익률 {pos["ret"]:+.1f}%\n'
            f'평가액 {pos["value"]:,}원\n\n'
            f'{debate}\n\n'
            f'<i>─ {pos["account"]} 계좌 ─</i>'
        )
        _send_telegram(msg)

    footer = (
        '\n<b>📋 CEO 판단 요청</b>\n'
        '각 종목의 [권고]에 대해 결정을 내려주세요.\n'
        '"1번 매수확대", "3번 축소" 형식으로 알려주시면 실행합니다.\n\n'
        '<i>* GM Capital AI 투자위원회 — 억지 토론 금지 원칙</i>'
    )
    _send_telegram(footer)
    print(f'✅ 투자위원회 리포트 발송 완료 ({len(positions)}종목)!')


if __name__ == '__main__':
    run_committee()
