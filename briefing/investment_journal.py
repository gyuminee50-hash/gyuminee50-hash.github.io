"""
GM Capital 투자일지 v5
시트: 미래에셋_거래(₩) / 메리츠_거래($) / 포지션×2 / 대시보드 / 분석히스토리
Excel SUMPRODUCT 수식이 포지션 자동계산 — Python은 현재가 + AI 분석만 담당
메리츠: 매입환율 컬럼 추가 → 환차손익·총수익(₩) 자동계산
"""
import json, os, re, time, threading, traceback
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import yfinance as yf
import requests

import groq_client

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
ONEDRIVE   = r'C:\Users\DeskTop\OneDrive\문서'
EXCEL_PATH = os.path.join(ONEDRIVE, 'GMCapital_투자일지.xlsx')

# ── 컬러 ─────────────────────────────────────────────────────────
W      = 'FFFFFF'
LIGHT  = 'F8FAFC'
H_NAVY = '1E3A5F'
H_SUB  = '243B53'
GOLD   = 'C9A84C'
AI_BG  = 'EBF4FF'
FB_BG  = 'FFF9EC'
G_BG   = 'DCFCE7'
R_BG   = 'FEE2E2'
G_TXT  = '15803D'
R_TXT  = 'DC2626'
DARK   = '1E293B'
BDR    = 'CBD5E1'

# ── 숫자 포맷 ─────────────────────────────────────────────────────
FMT_INT  = '#,##0'
FMT_DEC2 = '#,##0.00'
FMT_DEC4 = '#,##0.0000'
FMT_PCT  = '[Blue]+0.00"%";[Red]-0.00"%";"-"'
FMT_PNL  = '[Green]+#,##0.00;[Red]-#,##0.00;"-"'
FMT_PNLI = '[Green]+#,##0;[Red]-#,##0;"-"'

# ── 스타일 헬퍼 ──────────────────────────────────────────────────
def _fill(c):   return PatternFill('solid', fgColor=c)
def _font(bold=False, color=DARK, size=10):
    return Font(bold=bold, color=color, size=size, name='맑은 고딕')
def _align(h='center', wrap=True):
    return Alignment(horizontal=h, vertical='center', wrap_text=wrap)
def _border(c=BDR):
    s = Side(style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def _hborder():
    s = Side(style='medium', color='1A3050')
    return Border(left=s, right=s, top=s, bottom=s)

def _c(ws, r, col, val=None, bg=W, bold=False, color=DARK,
       fmt=None, h='center', wrap=True, size=10, hdr=False):
    cell = ws.cell(r, col)
    if val is not None:
        cell.value = val
    cell.fill      = _fill(bg)
    cell.font      = _font(bold=bold, color=color, size=size)
    cell.alignment = _align(h=h, wrap=wrap)
    cell.border    = _hborder() if hdr else _border()
    if fmt:
        cell.number_format = fmt
    return cell


# ══════════════════════════════════════════════════════════════════
# 거래 시트 (미래에셋_거래 / 메리츠_거래)
# ══════════════════════════════════════════════════════════════════
# 미래에셋: A날짜 B구분 C티커 D종목명 E가격(₩) F수량 G메모
# 메리츠:   A날짜 B구분 C티커 D종목명 E가격($) F수량 G매입환율 H메모
TRADE_COLS_KRW = [
    ('날짜',     12, None,    'center'),
    ('구분',      9, None,    'center'),
    ('티커',     11, None,    'center'),
    ('종목명',   22, None,    'center'),
    ('가격(₩)', 15, FMT_INT, 'center'),
    ('수량',      9, FMT_INT, 'center'),
    ('메모',     26, None,    'left'),
]
TRADE_COLS_USD = [
    ('날짜',      12, None,    'center'),
    ('구분',       9, None,    'center'),
    ('티커',      11, None,    'center'),
    ('종목명',    22, None,    'center'),
    ('가격($)',   15, FMT_DEC4,'center'),
    ('수량',       9, FMT_INT, 'center'),
    ('매입환율',  13, FMT_INT, 'center'),   # 매수 시점 USD/KRW 환율
    ('메모',      26, None,    'left'),
]

def _build_trade_sheet(wb, name, currency, tab, cols):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab
    last = get_column_letter(len(cols))
    # 서브타이틀 행 (행 1)
    ws.merge_cells(f'A1:{last}1')
    sub = ws['A1']
    sub.value = f'{name}  ({currency})'
    sub.fill  = _fill(H_NAVY); sub.font = _font(bold=True, color=GOLD, size=11)
    sub.alignment = _align(); sub.border = _hborder()
    ws.row_dimensions[1].height = 26
    # 헤더 (행 2)
    for ci, (label, width, fmt, h) in enumerate(cols, 1):
        _c(ws, 2, ci, label, H_SUB, bold=True, color=W, hdr=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = 'A3'
    # 데이터 영역 (행 3~502, 500행)
    for row in range(3, 503):
        bg = LIGHT if row % 2 == 1 else W
        for ci, (_, _, fmt, h) in enumerate(cols, 1):
            _c(ws, row, ci, bg=bg, fmt=fmt, h=h)
        ws.row_dimensions[row].height = 18


# ══════════════════════════════════════════════════════════════════
# 포지션 시트 — 헤더만 (데이터 행은 Python이 추가)
# ══════════════════════════════════════════════════════════════════
POS_COLS_KRW = [
    # (이름, 너비, 포맷, 정렬, 그룹)
    # 그룹: calc=수식자동계산, price=Python현재가, ai=Groq분석, judge=종합판단
    ('티커',          9,  None,     'center', 'calc'),
    ('종목명',        22, None,     'center', 'calc'),
    ('총매수수량',    10, FMT_INT,  'center', 'calc'),
    ('총매도수량',    10, FMT_INT,  'center', 'calc'),
    ('보유수량',      10, FMT_INT,  'center', 'calc'),
    ('평균단가(₩)',   14, FMT_INT,  'center', 'calc'),
    ('투자원금(₩)',   15, FMT_INT,  'center', 'calc'),
    ('현재가(₩)',     14, FMT_INT,  'center', 'price'),
    ('평가금액(₩)',   15, FMT_INT,  'center', 'price'),
    ('손익(₩)',       14, FMT_PNLI, 'center', 'price'),
    ('수익률(%)',     11, FMT_PCT,  'center', 'price'),
    ('매수 핵심 논거',52, None,     'left',   'ai'),
    ('리스크 요인',   32, None,     'left',   'ai'),
    ('목표가(₩)',     13, FMT_INT,  'center', 'ai'),
    ('손절가(₩)',     13, FMT_INT,  'center', 'ai'),
    ('종합판단',      12, None,     'center', 'judge'),
]
POS_COLS_USD = [
    ('티커',           9,  None,     'center', 'calc'),
    ('종목명',         22, None,     'center', 'calc'),
    ('총매수수량',     10, FMT_INT,  'center', 'calc'),
    ('총매도수량',     10, FMT_INT,  'center', 'calc'),
    ('보유수량',       10, FMT_INT,  'center', 'calc'),
    ('평균단가($)',    14, FMT_DEC4, 'center', 'calc'),
    ('투자원금($)',    15, FMT_DEC2, 'center', 'calc'),
    ('현재가($)',      14, FMT_DEC2, 'center', 'price'),
    ('평가금액($)',    15, FMT_DEC2, 'center', 'price'),
    ('손익($)',        14, FMT_PNL,  'center', 'price'),
    ('수익률(%)',      11, FMT_PCT,  'center', 'price'),
    ('매수 핵심 논거', 52, None,     'left',   'ai'),
    ('리스크 요인',    32, None,     'left',   'ai'),
    ('목표가($)',      13, FMT_DEC2, 'center', 'ai'),
    ('손절가($)',      13, FMT_DEC2, 'center', 'ai'),
    ('종합판단',       12, None,     'center', 'judge'),
    # ── 환차손익 전용 열 (Q~T) — 메리츠 전용, 수식 자동계산 ──
    ('매입환율(평균)', 13, FMT_INT,  'center', 'fx'),   # Q: SUMPRODUCT 가중평균
    ('환차손익(₩)',   14, FMT_PNLI, 'center', 'fx'),   # R: (현재환율-매입환율)×투자원금$
    ('총수익(₩)',     14, FMT_PNLI, 'center', 'fx'),   # S: 평가금액₩ - 투자원금₩
    ('투자원금(₩)',   15, FMT_INT,  'center', 'fx'),   # T: 투자원금$ × 매입환율
]

_HDR_BG = {
    'calc':  H_NAVY,   # 자동계산 열 — 어두운 네이비
    'price': '155E75', # 현재가 열 — 틸
    'ai':    H_SUB,    # AI 분석 열
    'judge': '3B1F6B', # 판단 열 — 퍼플
    'fx':    '065F46', # 환차손익 열 — 에메랄드 그린
}
_ROW_BG = {
    'calc':  None,   # 교대색 (W / LIGHT)
    'price': None,
    'ai':    AI_BG,
    'judge': FB_BG,
    'fx':    None,
}

def _build_position(wb, name, cols, tab):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'
    ws.sheet_properties.tabColor = tab
    for ci, (label, width, fmt, h, grp) in enumerate(cols, 1):
        _c(ws, 1, ci, label, _HDR_BG[grp], bold=True, color=GOLD, hdr=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 26


# ══════════════════════════════════════════════════════════════════
# 포지션 수식 생성
# ══════════════════════════════════════════════════════════════════
def _make_formulas(trade_sheet, row, is_us=False):
    """포지션 시트 row 번호에 해당하는 수식 딕셔너리 반환
    is_us=True면 환차손익 수식 Q~T열 추가 (메리츠_거래 G열 = 매입환율)
    """
    ts = f"'{trade_sheet}'"
    r  = row
    formulas = {
        # C: 총매수수량
        3:  (f"=SUMPRODUCT(({ts}!$C$3:$C$502=A{r})"
             f"*({ts}!$B$3:$B$502=\"매수\")"
             f"*({ts}!$F$3:$F$502))"),
        # D: 총매도수량
        4:  (f"=SUMPRODUCT(({ts}!$C$3:$C$502=A{r})"
             f"*({ts}!$B$3:$B$502=\"매도\")"
             f"*({ts}!$F$3:$F$502))"),
        # E: 보유수량
        5:  f"=C{r}-D{r}",
        # F: 가중평균단가 = Σ(가격×수량) / Σ(수량) — 매수 기준
        6:  (f"=IFERROR(SUMPRODUCT(({ts}!$C$3:$C$502=A{r})"
             f"*({ts}!$B$3:$B$502=\"매수\")"
             f"*{ts}!$E$3:$E$502"
             f"*{ts}!$F$3:$F$502)/C{r},0)"),
        # G: 투자원금($)
        7:  f"=F{r}*E{r}",
        # H: 현재가 — Python이 채움 (수식 없음)
        # I: 평가금액
        9:  f"=IF(H{r}>0,E{r}*H{r},0)",
        # J: 손익
        10: f"=IF(H{r}>0,I{r}-G{r},0)",
        # K: 수익률(%)
        11: f"=IF(AND(H{r}>0,G{r}>0),(I{r}-G{r})/G{r}*100,0)",
    }
    if is_us:
        # Q(17): 매입환율 가중평균 = Σ(가격×수량×환율) / Σ(가격×수량)
        #        메리츠_거래 G열 = 매입환율
        formulas[17] = (
            f"=IFERROR("
            f"SUMPRODUCT(({ts}!$C$3:$C$502=A{r})*({ts}!$B$3:$B$502=\"매수\")"
            f"*{ts}!$E$3:$E$502*{ts}!$F$3:$F$502*{ts}!$G$3:$G$502)"
            f"/"
            f"SUMPRODUCT(({ts}!$C$3:$C$502=A{r})*({ts}!$B$3:$B$502=\"매수\")"
            f"*{ts}!$E$3:$E$502*{ts}!$F$3:$F$502)"
            f",0)"
        )
        # R(18): 환차손익(₩) = (현재환율 - 매입환율) × 투자원금($)
        formulas[18] = f"=IF(Q{r}>0,('대시보드'!$U$2-Q{r})*G{r},0)"
        # S(19): 총수익(₩) = 평가금액($)×현재환율 - 투자원금($)×매입환율
        formulas[19] = f"=IF(AND(I{r}>0,Q{r}>0),I{r}*'대시보드'!$U$2-G{r}*Q{r},0)"
        # T(20): 투자원금(₩) = 투자원금($) × 매입환율
        formulas[20] = f"=IF(Q{r}>0,G{r}*Q{r},0)"
    return formulas


def _add_pos_row(ws, row, ticker, name, trade_sheet, cols, is_us=False):
    """포지션 시트에 새 티커 행 추가: A·B=값, C~K=수식, L~P=빈칸(AI담당)
    is_us=True 이면 Q~T(환차손익) 수식도 추가
    """
    formulas = _make_formulas(trade_sheet, row, is_us=is_us)
    alt_bg   = LIGHT if row % 2 == 0 else W

    for ci, (_, _, fmt, h, grp) in enumerate(cols, 1):
        row_bg = _ROW_BG[grp] or alt_bg
        if ci == 1:
            val = ticker
        elif ci == 2:
            val = name
        elif ci in formulas:
            val = formulas[ci]
        else:
            val = None   # H(현재가) 및 L~P(AI분석) — Python이 별도 채움

        _c(ws, row, ci, val, row_bg, fmt=fmt, h=h, wrap=(ci >= 12))
    ws.row_dimensions[row].height = 70


# ══════════════════════════════════════════════════════════════════
# 대시보드  (v5 — 3구역 재설계)
# ══════════════════════════════════════════════════════════════════
# 레이아웃:
#   Row 1      : 타이틀
#   Row 2      : 환율 (R2:T2 레이블 / U2 값 — Python 업데이트)
#   Row 3      : 미래에셋 섹션 헤더 A3:K3  |  메리츠 섹션 헤더 M3:U3
#   Row 4      : 카드 레이블 (미래에셋 4개 + 메리츠 4개)
#   Row 5      : 카드 값 (큰 폰트)
#   Row 6      : 전체 합계 섹션 헤더
#   Row 7      : 전체 합계 레이블
#   Row 8      : 전체 합계 값
#   Row 10     : 보유 현황 헤더 (미래에셋 + 메리츠)
#   Row 11     : 컬럼 헤더
#   Row 12~    : 데이터 (포지션 시트 참조)
# ══════════════════════════════════════════════════════════════════
def _card(ws, lbl_rng, val_rng, label, formula, fmt,
          lbl_bg=H_NAVY, val_bg=W, lbl_color=GOLD, val_color=H_NAVY):
    """레이블+값 카드 한 쌍 생성"""
    for rng in [lbl_rng, val_rng]:
        if ':' in rng:
            ws.merge_cells(rng)
    lc = ws[lbl_rng.split(':')[0]]
    lc.value = label
    lc.fill  = _fill(lbl_bg); lc.font = _font(bold=True, color=lbl_color, size=9)
    lc.alignment = _align(); lc.border = _hborder()

    vc = ws[val_rng.split(':')[0]]
    vc.value = formula
    vc.fill  = _fill(val_bg)
    vc.font  = Font(bold=True, color=val_color, size=13, name='맑은 고딕')
    vc.alignment = _align()
    vc.border = _hborder()
    vc.number_format = fmt


def _build_dashboard(wb):
    ws = wb.create_sheet('대시보드')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '10B981'

    for r in range(1, 65):
        for c in range(1, 22):
            ws.cell(r, c).fill = _fill(LIGHT)

    # ── Row 1: 타이틀 ──
    ws.merge_cells('A1:U1')
    t = ws['A1']
    t.value = 'GM Capital  Investment Dashboard'
    t.fill  = _fill(H_NAVY)
    t.font  = Font(bold=True, color=GOLD, size=15, name='맑은 고딕')
    t.alignment = _align()
    ws.row_dimensions[1].height = 34

    # ── Row 2: 환율 — Python이 U2를 업데이트, 모든 수식이 U2 참조 ──
    ws.merge_cells('R2:T2')
    _c(ws, 2, 18, 'USD/KRW 환율', H_SUB, bold=True, color=W, hdr=True)
    _c(ws, 2, 21, 1380, W, bold=True, color=H_NAVY, fmt=FMT_INT, size=12)
    ws.row_dimensions[2].height = 24

    # ── Row 3: 섹션 헤더 ──
    ws.merge_cells('A3:K3')
    h = ws['A3']
    h.value = '  미래에셋  (국내 · 원화)'
    h.fill  = _fill(H_NAVY); h.font = _font(bold=True, color=GOLD, size=11)
    h.alignment = _align(h='left'); h.border = _hborder()

    ws.merge_cells('M3:U3')
    h2 = ws['M3']
    h2.value = '  메리츠  (해외 · 달러)'
    h2.fill  = _fill('1D4ED8'); h2.font = _font(bold=True, color=W, size=11)
    h2.alignment = _align(h='left'); h2.border = _hborder()
    ws.row_dimensions[3].height = 22

    # ── Row 4~5: 카드 (레이블 + 값) ──
    # 미래에셋 4개 카드: A~C | D~F | G~I | J~K
    dom_inv = "SUM('미래에셋_포지션'!G:G)"
    _card(ws, 'A4:C4', 'A5:C5', '투자원금(₩)',
          f'={dom_inv}', FMT_INT)
    _card(ws, 'D4:F4', 'D5:F5', '평가금액(₩)',
          "=SUM('미래에셋_포지션'!I:I)", FMT_INT)
    _card(ws, 'G4:I4', 'G5:I5', '매매손익(₩)',
          "=SUM('미래에셋_포지션'!J:J)", FMT_PNLI)
    _card(ws, 'J4:K4', 'J5:K5', '수익률(%)',
          f"=IFERROR(SUM('미래에셋_포지션'!J:J)/{dom_inv}*100,0)", FMT_PCT)

    # 메리츠 4개 카드: M~N | O~P | Q~S | T~U
    # (매매손익은 $ × 현재환율, 환차손익·총수익은 포지션 수식이 직접 계산)
    _card(ws, 'M4:N4', 'M5:N5', '매매손익(₩)',
          "=SUM('메리츠_포지션'!J:J)*U2", FMT_PNLI,
          lbl_bg='1E3A8A', val_color='1D4ED8')
    _card(ws, 'O4:P4', 'O5:P5', '환차손익(₩)',
          "=SUM('메리츠_포지션'!R:R)", FMT_PNLI,
          lbl_bg='065F46', val_color='065F46')
    _card(ws, 'Q4:S4', 'Q5:S5', '총수익(₩)',
          "=SUM('메리츠_포지션'!S:S)", FMT_PNLI,
          lbl_bg='1D4ED8', val_color='1D4ED8')
    _card(ws, 'T4:U4', 'T5:U5', '수익률(환포함)',
          "=IFERROR(SUM('메리츠_포지션'!S:S)/SUM('메리츠_포지션'!T:T)*100,0)", FMT_PCT,
          lbl_bg='1E3A8A', val_color='1D4ED8')

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 34

    # ── Row 6~8: 전체 합계 ──
    ws.merge_cells('A6:U6')
    g = ws['A6']
    g.value = '  전체 합계'
    g.fill  = _fill('0F4C2A'); g.font = _font(bold=True, color='DCFCE7', size=10)
    g.alignment = _align(h='left'); g.border = _hborder()
    ws.row_dimensions[6].height = 18

    tot_inv = "SUM('미래에셋_포지션'!G:G)+SUM('메리츠_포지션'!T:T)"
    tot_pnl = "SUM('미래에셋_포지션'!J:J)+SUM('메리츠_포지션'!S:S)"
    _card(ws, 'A7:F7',  'A8:F8',  '전체 투자원금(₩)',
          f'={tot_inv}', FMT_INT,
          lbl_bg='14532D', lbl_color='BBF7D0', val_bg='F0FDF4', val_color='166534')
    _card(ws, 'G7:N7',  'G8:N8',  '전체 총손익(₩)',
          f'={tot_pnl}', FMT_PNLI,
          lbl_bg='14532D', lbl_color='BBF7D0', val_bg='F0FDF4', val_color='166534')
    _card(ws, 'O7:U7',  'O8:U8',  '전체 수익률(%)',
          f'=IFERROR(({tot_pnl})/({tot_inv})*100,0)', FMT_PCT,
          lbl_bg='14532D', lbl_color='BBF7D0', val_bg='F0FDF4', val_color='166534')

    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 36

    # ── Row 10~: 보유 현황 테이블 ──
    RH = 10  # holdings header row

    # 미래에셋 보유 현황
    ws.merge_cells(f'A{RH}:G{RH}')
    bh = ws[f'A{RH}']
    bh.value = '  미래에셋 보유 현황 (국내 · 원화)'
    bh.fill  = _fill(H_NAVY); bh.font = _font(bold=True, color=W, size=11)
    bh.alignment = _align(h='left')
    ws.row_dimensions[RH].height = 24

    dom_h    = ['티커','종목명','보유수량','평균단가(₩)','현재가(₩)','손익(₩)','수익률(%)']
    dom_fmts = [None, None, FMT_INT, FMT_INT, FMT_INT, FMT_PNLI, FMT_PCT]
    dom_col  = {1:'A', 2:'B', 3:'E', 4:'F', 5:'H', 6:'J', 7:'K'}
    for ci, (h, f) in enumerate(zip(dom_h, dom_fmts), 1):
        _c(ws, RH+1, ci, h, H_SUB, bold=True, color=W, fmt=f, hdr=True)
    ws.row_dimensions[RH+1].height = 20

    for r in range(RH+2, RH+14):   # 12행 = 최대 12종목
        bg = W if r % 2 == 0 else LIGHT
        pr = r - RH
        for ci, fmt in enumerate(dom_fmts, 1):
            _c(ws, r, ci, f"='미래에셋_포지션'!{dom_col[ci]}{pr}", bg, fmt=fmt, size=10)
        ws.row_dimensions[r].height = 20

    # 메리츠 보유 현황
    ws.merge_cells(f'I{RH}:P{RH}')
    bh2 = ws[f'I{RH}']
    bh2.value = '  메리츠 보유 현황 (해외 · 달러)'
    bh2.fill  = _fill('1D4ED8'); bh2.font = _font(bold=True, color=W, size=11)
    bh2.alignment = _align(h='left')

    us_h    = ['티커','종목명','보유수량','평균단가($)','현재가($)','손익($)','수익률(%)']
    us_fmts = [None, None, FMT_INT, FMT_DEC4, FMT_DEC2, FMT_PNL, FMT_PCT]
    us_col  = {1:'A', 2:'B', 3:'E', 4:'F', 5:'H', 6:'J', 7:'K'}
    for ci, (h, f) in enumerate(zip(us_h, us_fmts), 9):
        _c(ws, RH+1, ci, h, '1E40AF', bold=True, color=W, fmt=f, hdr=True)

    for r in range(RH+2, RH+14):
        bg = W if r % 2 == 0 else 'EFF6FF'
        pr = r - RH
        for ci, fmt in enumerate(us_fmts, 1):
            col = us_col[ci]
            _c(ws, r, ci+8, f"='메리츠_포지션'!{col}{pr}", bg, fmt=fmt, size=10)
        ws.row_dimensions[r].height = 20

    # ── 컬럼 너비 ──
    for col, w in [
        ('A',9),('B',22),('C',12),('D',14),('E',14),('F',14),('G',12),
        ('H',2), ('I',9),('J',22),('K',12),('L',12),('M',12),('N',12),
        ('O',13),('P',13),('Q',13),('R',13),('S',13),('T',13),('U',12),
    ]:
        ws.column_dimensions[col].width = w


def _build_history(wb):
    ws = wb.create_sheet('분석히스토리')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '94A3B8'
    cols = [('분석일시',18),('계좌',12),('티커',10),('구분',12),('핵심논거',80)]
    for ci, (n, w) in enumerate(cols, 1):
        _c(ws, 1, ci, n, H_NAVY, bold=True, color=W, hdr=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    for r in range(2, 201):
        bg = LIGHT if r % 2 == 0 else W
        for ci in range(1, 6):
            _c(ws, r, ci, bg=bg, size=9, h='left' if ci == 5 else 'center')


# ══════════════════════════════════════════════════════════════════
# 초기 보유 데이터
# ══════════════════════════════════════════════════════════════════
def _init_holdings(wb):
    # 이미 데이터 있으면 건너뜀
    if wb['미래에셋_거래'].cell(3, 1).value:
        return

    today = '2026-06-07'
    dom_rows = [
        # 날짜, 구분, 티커, 종목명, 가격(₩), 수량, 메모
        (today, '매수', '133690', 'TIGER 미국나스닥100',    191952,   8, '초기보유'),
        (today, '매수', '360750', 'TIGER 미국S&P500',        25481, 166, '초기보유'),
        (today, '매수', '232080', 'TIGER 코스닥150',         17668, 220, '초기보유'),
        (today, '매수', '453850', 'ACE 미국30년국채액티브',   7908, 372, '초기보유'),
    ]
    # 메리츠_거래: 8컬럼 — 날짜,구분,티커,종목명,가격($),수량,매입환율,메모
    us_rows = [
        (today, '매수', 'MU',   'Micron Technology',        708.46,   2, 1501.5,  '초기보유'),
        (today, '매수', 'IEMG', 'iShares Core MSCI EM',    70.5376,  13, 1467.48, '초기보유'),
        (today, '매수', 'SPYM', 'SPDR Portfolio S&P500',   80.7258,  24, 1464.67, '초기보유'),
        (today, '매수', 'GGLL', 'GraniteShares 2x GOOGL', 113.7276,  17, 1492.42, '초기보유'),
        (today, '매수', 'QLD',  'ProShares Ultra QQQ 2X',  85.0420,  24, 1463.62, '초기보유'),
        (today, '매수', 'TSM',  'TSMC ADR',                406.7400,   1, 1499.79, '초기보유'),
    ]
    fmts_dom = [None, None, None, None, FMT_INT, FMT_INT, None]           # 7 cols
    fmts_us  = [None, None, None, None, FMT_DEC4, FMT_INT, FMT_INT, None] # 8 cols

    ws_dom = wb['미래에셋_거래']
    for i, rd in enumerate(dom_rows, 3):
        bg = LIGHT if i % 2 == 1 else W
        for ci, val in enumerate(rd, 1):
            _c(ws_dom, i, ci, val, bg, fmt=fmts_dom[ci-1])

    ws_us = wb['메리츠_거래']
    for i, rd in enumerate(us_rows, 3):
        bg = LIGHT if i % 2 == 1 else W
        for ci, val in enumerate(rd, 1):
            _c(ws_us, i, ci, val, bg, fmt=fmts_us[ci-1])


# ══════════════════════════════════════════════════════════════════
# 현재가 조회
# ══════════════════════════════════════════════════════════════════
def _usd_krw():
    try:
        h = yf.Ticker('USDKRW=X').history(period='2d')
        return round(float(h['Close'].iloc[-1]), 0) if not h.empty else 1380.0
    except Exception:
        return 1380.0

def _kr_price(ticker):
    try:
        h = yf.Ticker(f'{str(ticker).zfill(6)}.KS').history(period='2d')
        return int(round(float(h['Close'].iloc[-1]))) if not h.empty else 0
    except Exception:
        return 0

def _us_prices(tickers):
    if not tickers:
        return {}
    try:
        data = yf.download(list(tickers), period='2d', progress=False, auto_adjust=True)
        out  = {}
        for t in tickers:
            try:
                col = data['Close'][t] if len(tickers) > 1 else data['Close']
                out[t] = round(float(col.dropna().iloc[-1]), 4)
            except Exception:
                out[t] = 0.0
        return out
    except Exception:
        return {t: 0.0 for t in tickers}


# ══════════════════════════════════════════════════════════════════
# 거래 데이터 파싱 (AI 프롬프트용 내부 계산)
# ══════════════════════════════════════════════════════════════════
def _parse_tickers(wb, trade_sheet):
    """거래 시트에서 {ticker: name} 순서 보존 반환"""
    ws   = wb[trade_sheet]
    seen = {}
    for r in range(3, ws.max_row + 1):
        t = ws.cell(r, 3).value   # C열 = 티커
        n = ws.cell(r, 4).value   # D열 = 종목명
        if t and t not in seen:
            seen[t] = n or t
    return seen

def _calc_avg(wb, trade_sheet, ticker):
    """가중평균단가 + 보유수량 계산 (Groq 프롬프트 구성용)"""
    ws         = wb[trade_sheet]
    total_cost = total_qty = 0
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 3).value != ticker:
            continue
        kind  = str(ws.cell(r, 2).value or '')
        price = ws.cell(r, 5).value or 0
        qty   = ws.cell(r, 6).value or 0
        if '매수' in kind:
            total_cost += price * qty
            total_qty  += qty
        elif '매도' in kind:
            total_qty  -= qty
    hold = max(0, total_qty)
    avg  = (total_cost / (total_cost / hold if hold else 1)) if hold else 0
    # 더 단순하게
    buy_cost = buy_qty = 0
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 3).value != ticker:
            continue
        if '매수' in str(ws.cell(r, 2).value or ''):
            p = ws.cell(r, 5).value or 0
            q = ws.cell(r, 6).value or 0
            buy_cost += p * q
            buy_qty  += q
    avg = round(buy_cost / buy_qty, 4) if buy_qty else 0
    return avg, hold


# ══════════════════════════════════════════════════════════════════
# Groq 분석
# ══════════════════════════════════════════════════════════════════
def _fetch_news(ticker):
    try:
        import xml.etree.ElementTree as ET
        url  = (f'https://news.google.com/rss/search?q={ticker}+stock'
                '&hl=en-US&gl=US&ceid=US:en')
        resp = requests.get(url, timeout=8, headers={'User-Agent': 'Mozilla/5.0'})
        items = ET.fromstring(resp.content).findall('.//item')
        return ' / '.join(
            i.find('title').text.strip() for i in items[:3]
            if i.find('title') is not None
        ) or '뉴스 없음'
    except Exception:
        return '뉴스 없음'

def _fetch_info(ticker):
    try:
        info = yf.Ticker(ticker).info
        return {
            'target': round(info.get('targetMeanPrice') or 0, 2),
            'rec':    info.get('recommendationKey', ''),
            'pe':     round(info.get('trailingPE') or 0, 1),
            'sector': info.get('sector', ''),
        }
    except Exception:
        return {}

_DOM = """\
국내 ETF 보유 분석.
종목: {name} ({ticker}) | 평균단가: {price:,}원 | 보유: {qty}주 | 투자원금: {inv:,}원
현재가: {cur:,}원 | 수익률: {ret:+.1f}%

JSON만 응답:
{{"핵심논거":"보유 유지 이유 (매크로·섹터·ETF 특성 근거 3가지, 250자 이내)","리스크":"1. 리스크1\\n2. 리스크2","목표가":숫자,"손절가":숫자,"종합판단":"매수적절 or 주의 or 부적절"}}"""

_US = """\
미국 주식/ETF 보유 분석.
종목: {name} ({ticker}) | 평균단가: ${price} | 보유: {qty}주 | 투자원금: ${inv:,.0f}
현재가: ${cur} | 수익률: {ret:+.1f}% | 목표주가: ${target} | 추천: {rec} | PER: {pe} | 섹터: {sector}
뉴스: {news}

JSON만 응답:
{{"핵심논거":"보유 유지 이유 (펀더멘털·매크로·밸류에이션 근거 3가지+, 280자 이내)","리스크":"1. 리스크1\\n2. 리스크2","목표가":숫자,"손절가":숫자,"종합판단":"매수적절 or 주의 or 부적절"}}"""

def _run_analysis(is_us, ticker, name, avg, qty, cur):
    ret = (cur - avg) / avg * 100 if avg else 0
    if is_us:
        info   = _fetch_info(ticker)
        news   = _fetch_news(ticker)
        prompt = _US.format(name=name, ticker=ticker, price=avg, qty=qty,
                            inv=avg*qty, cur=cur, ret=ret,
                            target=info.get('target',0), rec=info.get('rec',''),
                            pe=info.get('pe',0), sector=info.get('sector',''),
                            news=news)
    else:
        prompt = _DOM.format(name=name, ticker=ticker,
                             price=int(avg), qty=qty, inv=int(avg*qty),
                             cur=int(cur), ret=ret)
    try:
        raw = groq_client.call(prompt, max_tokens=700, temperature=0.4)
        m   = re.search(r'\{.*\}', raw, re.DOTALL)
        return json.loads(m.group()) if m else {}
    except Exception as e:
        print(f'    [분석오류 {ticker}] {e}')
        return {}


# ══════════════════════════════════════════════════════════════════
# 포지션 시트 업데이트 (Python이 하는 것: 신규행 추가 + 현재가 + AI분석)
# ══════════════════════════════════════════════════════════════════
def _existing_tickers(ws):
    """포지션 시트의 {ticker: row} 반환"""
    out = {}
    for r in range(2, ws.max_row + 1):
        t = ws.cell(r, 1).value
        if t:
            out[t] = r
    return out

def _write_analysis(ws, row, result, is_us):
    for ci, key in [(12,'핵심논거'),(13,'리스크'),(14,'목표가'),(15,'손절가'),(16,'종합판단')]:
        val = result.get(key, '')
        bg  = AI_BG if ci <= 15 else FB_BG
        fmt = (FMT_DEC2 if is_us else FMT_INT) if ci in (14, 15) else None
        _c(ws, row, ci, val, bg, fmt=fmt, h='left' if ci <= 13 else 'center', wrap=True)
    # 종합판단 컬러 강조
    val   = str(result.get('종합판단', ''))
    color = G_TXT if val == '매수적절' else (R_TXT if val == '부적절' else 'B45309')
    bg    = G_BG  if val == '매수적절' else (R_BG  if val == '부적절' else FB_BG)
    c = ws.cell(row, 16)
    c.fill = _fill(bg); c.font = _font(bold=True, color=color)

def update_positions(wb, trade_sheet, pos_sheet, is_us, prices):
    ws_pos    = wb[pos_sheet]
    cols      = POS_COLS_USD if is_us else POS_COLS_KRW
    in_trade  = _parse_tickers(wb, trade_sheet)
    in_sheet  = _existing_tickers(ws_pos)

    for ticker, name in in_trade.items():
        # ① 신규 티커 → 수식 행 추가
        if ticker not in in_sheet:
            next_row = max(in_sheet.values(), default=1) + 1
            _add_pos_row(ws_pos, next_row, ticker, name, trade_sheet, cols, is_us=is_us)
            in_sheet[ticker] = next_row
            print(f'    [신규] {ticker} → row {next_row}')

        row = in_sheet[ticker]

        # ② 현재가(H열) 업데이트 — Python만 쓰는 유일한 숫자 열
        cur = (_kr_price(ticker) if not is_us else prices.get(ticker, 0))
        if cur:
            c = ws_pos.cell(row, 8)
            c.value          = cur
            c.number_format  = FMT_DEC2 if is_us else FMT_INT
            c.fill           = _fill(W)
            c.font           = _font(bold=True, color='155E75')
            c.alignment      = _align()
            c.border         = _border()

        # ③ AI 분석(L~P열) — 없는 경우만 실행
        if not ws_pos.cell(row, 12).value:
            avg, qty = _calc_avg(wb, trade_sheet, ticker)
            cur_or_avg = cur or avg
            print(f'    [분석] {ticker}')
            result = _run_analysis(is_us, ticker, name, avg, qty, cur_or_avg)
            if result:
                _write_analysis(ws_pos, row, result, is_us)
                _write_history(wb, pos_sheet, ticker, result.get('핵심논거', ''))
        else:
            print(f'    [캐시] {ticker}')


def update_dashboard_rate(wb, rate):
    """대시보드 환율 셀(U2) 업데이트"""
    wb['대시보드']['U2'].value = rate


def _write_history(wb, sheet, ticker, content):
    ws  = wb['분석히스토리']
    row = 2
    while ws.cell(row, 1).value:
        row += 1
    data = [datetime.now().strftime('%Y-%m-%d %H:%M'), sheet, ticker, '포지션분석', content]
    bg   = LIGHT if row % 2 == 0 else W
    for ci, val in enumerate(data, 1):
        _c(ws, row, ci, val, bg, size=9, h='left' if ci == 5 else 'center')


# ══════════════════════════════════════════════════════════════════
# 전체 처리
# ══════════════════════════════════════════════════════════════════
_lock = threading.Lock()

def process_all():
    if not os.path.exists(EXCEL_PATH):
        return

    with _lock:
        try:
            wb   = openpyxl.load_workbook(EXCEL_PATH)
            rate = _usd_krw()
            print(f'  환율: {rate:,.0f}원')

            dom_tickers = list(_parse_tickers(wb, '미래에셋_거래').keys())
            us_tickers  = list(_parse_tickers(wb, '메리츠_거래').keys())
            print(f'  미래에셋 {len(dom_tickers)}종목 / 메리츠 {len(us_tickers)}종목')

            prices = _us_prices(us_tickers)

            update_positions(wb, '미래에셋_거래', '미래에셋_포지션', False, {})
            update_positions(wb, '메리츠_거래',   '메리츠_포지션',   True,  prices)
            update_dashboard_rate(wb, rate)

            wb.save(EXCEL_PATH)
            print(f'  저장완료 ({datetime.now().strftime("%H:%M:%S")})')

        except Exception as e:
            print(f'  [오류] {e}')
            traceback.print_exc()


# ══════════════════════════════════════════════════════════════════
# Watchdog + 주기 업데이트
# ══════════════════════════════════════════════════════════════════
class ExcelHandler(FileSystemEventHandler):
    def __init__(self):
        self._last = 0
    def on_modified(self, event):
        if 'GMCapital_투자일지' in event.src_path:
            now = time.time()
            if now - self._last < 8:
                return
            self._last = now
            print(f'\n[감지] 파일변경 → 재처리')
            threading.Thread(target=process_all, daemon=True).start()

def _price_loop():
    while True:
        time.sleep(3600)
        print(f'\n[{datetime.now().strftime("%H:%M")}] 정기 가격 업데이트')
        process_all()


def main():
    needs_create = True
    if os.path.exists(EXCEL_PATH):
        try:
            wbc = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
            needs_create = '미래에셋_거래' not in wbc.sheetnames
            wbc.close()
        except Exception:
            pass

    if needs_create:
        print('엑셀 새로 생성 중...')
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _build_trade_sheet(wb, '미래에셋_거래', '원화(₩)', 'C9A84C', TRADE_COLS_KRW)
        _build_trade_sheet(wb, '메리츠_거래',   '달러($)',  '3B82F6', TRADE_COLS_USD)
        _build_position(wb, '미래에셋_포지션', POS_COLS_KRW, 'B45309')
        _build_position(wb, '메리츠_포지션',   POS_COLS_USD, '1D4ED8')
        _build_dashboard(wb)
        _build_history(wb)
        _init_holdings(wb)
        wb.save(EXCEL_PATH)
        print(f'  완료: {EXCEL_PATH}')

    print('초기 처리 중...')
    process_all()

    threading.Thread(target=_price_loop, daemon=True).start()
    observer = Observer()
    observer.schedule(ExcelHandler(), path=ONEDRIVE, recursive=False)
    observer.start()
    print('감시 시작. 종료: Ctrl+C')
    try:
        while True:
            time.sleep(10)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()


if __name__ == '__main__':
    main()
